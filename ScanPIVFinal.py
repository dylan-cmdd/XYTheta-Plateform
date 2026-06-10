"""
============================================================================
  SCAN PIV  —  pilotage plateforme + camera Phantom   (version 6)
============================================================================

Ce que fait cette version, en plus du pilotage de base :
  - CONFIRMATION D'ARRIVEE : apres un deplacement, Python ne fait pas
    confiance au "c'est fini" renvoye par l'Arduino. Il relit la position
    en boucle et n'avance que lorsque la plateforme est SUR la cible ET
    immobile. Evite que le 1er enregistrement d'un balayage commence avant
    que la plateforme soit en place.
  - STABILISATION : petite pause apres l'arrivee, pour laisser retomber les
    vibrations mecaniques avant de filmer.
  - ATTENTE FIN D'ECRITURE : filmer() ne rend la main qu'une fois le .cine
    reellement ecrit sur le disque (surveillance de la taille du fichier).
  - JOURNAL EXCEL : pendant le scan on note les parametres de chaque film
    (position, resolution, fps, exposition...). A la fin (ou meme si le scan
    s'arrete en cours), on ecrit un recapitulatif .xlsx dans le dossier.

  -> Le recapitulatif Excel necessite la librairie 'openpyxl' :
        pip install openpyxl
     Sinon, le programme ecrit un .csv a la place (Excel l'ouvre aussi).

Outils de base :
  lire_position(arduino)                  ->  lit ou est la plateforme (X, Y, T)
  attendre_arrivee(arduino, x, y)         ->  attend que la plateforme soit en place
  aller_a(arduino, x, y)                  ->  va aux coordonnees absolues (x, y) + confirme
  bouger(arduino, direction, nb_pas)      ->  deplace d'un nombre de pas
  filmer(cam, nom, nb_images, fps)        ->  enregistre un film ET attend la fin d'ecriture
  ecrire_journal_excel(lignes, chemin)    ->  ecrit le recapitulatif des prises

Deroulement de la PIV :
  1) RESET (mise a l'origine 0,0,0) + confirmation
  2) MODE LARGEUR  : aller a (x_debut, y) face nappe, balayer X decroissant
  3) rotation de 90 degres
  4) MODE LONGUEUR : aller a (x_debut, y) face nappe, balayer X decroissant

Le declenchement camera se fait sur le front du creneau GBF (trigger
externe) : on ARME avec record(), c'est le front qui declenche. On
n'appelle donc jamais cam.trigger().
----------------------------------------------------------------------------
"""

import os
import time
import csv
import serial
import serial.tools.list_ports

from pyphantom import Phantom, utils
from Camera_Selector_Fn import camera_selector


# ============================================================
#   REGLAGES FIXES
# ============================================================

# Dossier ou enregistrer les films. A changer selon la machine.
DOSSIER_SORTIE = r"C:\Users\Dremaro\Desktop\StageDylan\WS_python\Film_PIV"

# Mode de synchro : None = detection auto. Sinon nom exact ("EXTERNAL"/"INTERNAL").
SYNC_MODE_NOM = "INTERNAL"

# Pause de stabilisation mecanique (s) apres l'arrivee a une position,
# avant de filmer. A augmenter si la 1ere image d'un balayage est floue.
STABILISATION_S = 0.5


# ============================================================
#   PETITES FONCTIONS POUR POSER LES QUESTIONS
# ============================================================

def demander_entier(message, defaut):
    """Pose une question et renvoie un entier. Entree seule => valeur par defaut."""
    while True:
        texte = input(f"{message} [{defaut}] : ").strip()
        if texte == "":
            return defaut
        try:
            return int(texte)
        except ValueError:
            print("  -> tape un nombre entier (ou Entree pour la valeur par defaut)")


def demander_decimal(message, defaut):
    """Pose une question et renvoie un nombre a virgule. Entree seule => defaut."""
    while True:
        texte = input(f"{message} [{defaut}] : ").strip()
        if texte == "":
            return defaut
        try:
            return float(texte)
        except ValueError:
            print("  -> tape un nombre (ou Entree pour la valeur par defaut)")


# ============================================================
#   MODE DE SYNCHRO
# ============================================================

def trouver_mode_sync():
    """Renvoie le mode de synchro a appliquer (cherche a l'execution)."""
    noms = [n for n in dir(utils.SyncModeEnum) if not n.startswith("_")]
    print("Modes de synchro disponibles dans le SDK :", noms)

    if SYNC_MODE_NOM is not None:
        return getattr(utils.SyncModeEnum, SYNC_MODE_NOM)

    for nom in noms:
        if "EXT" in nom.upper():
            print("  -> mode externe detecte :", nom)
            return getattr(utils.SyncModeEnum, nom)

    print("  ATTENTION : pas de mode 'externe' trouve, mode par defaut garde.")
    return None


# ============================================================
#   PLATEFORME : envoyer, lire la position, bouger, aller a
# ============================================================

def envoyer(arduino, commande):
    """Envoie une commande a l'Arduino et attend sa reponse (= mouvement fini).
    L'Arduino renvoie sa position apres CHAQUE commande recue."""

    # 1. On vide tous les vieux dechets (ex: codes IR) juste avant de parler
    arduino.reset_input_buffer()

    # 2. On envoie l'ordre
    arduino.write((commande + "\n").encode("utf-8"))

    # 3. On attend patiemment la reponse (marche uniquement si timeout=None)
    reponse = arduino.readline().decode("utf-8").strip()
    print(f"  reponse : {reponse}")

    return reponse


def lire_position(arduino):
    """Demande sa position a l'Arduino et la renvoie sous forme de dict.
    On envoie une commande neutre ('POS', non reconnue) : l'Arduino ne
    bouge pas, mais renvoie quand meme sa position.
    Reponse attendue de la forme : 'X : 120 Y : 4150 T : 0'."""
    reponse = envoyer(arduino, "POS")
    try:
        # On remplace les ':' par des espaces puis on decoupe :
        # "X : 120 Y : 4150 T : 0" -> ['X','120','Y','4150','T','0']
        morceaux = reponse.replace(":", " ").split()
        x = int(morceaux[1])
        y = int(morceaux[3])
        t = int(morceaux[5])
        return {"X": x, "Y": y, "T": t}
    except Exception:
        print("  ATTENTION : position illisible :", reponse)
        return {"X": 0, "Y": 0, "T": 0}


def bouger(arduino, direction, nb_pas):
    """Bouge la plateforme d'un nombre de pas dans une direction.
        direction : "XP", "XM", "YP", "YM", "TP" ou "TM"
    On regle d'abord la taille du pas (PAS<n>), puis on declenche le mouvement."""
    if nb_pas <= 0:
        return
    envoyer(arduino, "PAS" + str(nb_pas))
    envoyer(arduino, direction)
    print(f"  bouge {direction} de {nb_pas} pas")


def attendre_arrivee(arduino, x_cible, y_cible, tolerance=2, timeout=60):
    """Verifie que la plateforme est REELLEMENT arrivee a (x_cible, y_cible).
    On ne fait PAS confiance au 'c'est fini' renvoye par l'Arduino : on relit
    la position en boucle, et on ne valide que lorsqu'elle est sur la cible
    (a 'tolerance' pas pres) ET qu'elle ne bouge plus (deux lectures identiques).
    Renvoie True si l'arrivee est confirmee, False si on sort par timeout."""
    debut = time.time()
    precedente = None

    while time.time() - debut < timeout:
        pos = lire_position(arduino)

        sur_cible = (abs(pos["X"] - x_cible) <= tolerance and
                     abs(pos["Y"] - y_cible) <= tolerance)
        immobile = (precedente is not None and
                    pos["X"] == precedente["X"] and pos["Y"] == precedente["Y"])

        if sur_cible and immobile:
            print(f"  plateforme en place : X={pos['X']} Y={pos['Y']}")
            return True

        precedente = pos
        time.sleep(0.1)

    print(f"  ATTENTION : arrivee NON confirmee (cible X={x_cible} Y={y_cible}).")
    return False


def aller_a(arduino, x_cible, y_cible):
    """Va aux coordonnees ABSOLUES (x_cible, y_cible), puis ATTEND la confirmation
    que la plateforme y est vraiment avant de rendre la main.
    On lit la position actuelle, on calcule l'ecart, et on bouge dans le bon sens."""
    pos = lire_position(arduino)
    print(f"  position actuelle : X={pos['X']} Y={pos['Y']} -> cible X={x_cible} Y={y_cible}")

    # Deplacement en X
    dx = x_cible - pos["X"]
    if dx > 0:
        bouger(arduino, "XP", dx)
    elif dx < 0:
        bouger(arduino, "XM", -dx)   # -dx est positif ici

    # Deplacement en Y
    dy = y_cible - pos["Y"]
    if dy > 0:
        bouger(arduino, "YP", dy)
    elif dy < 0:
        bouger(arduino, "YM", -dy)

    # On VERIFIE l'arrivee (au lieu de croire l'Arduino sur parole) AVANT de continuer.
    attendre_arrivee(arduino, x_cible, y_cible)

    # Pause finale : stabilisation mecanique (vibrations residuelles apres le dernier pas).
    time.sleep(STABILISATION_S)


# ============================================================
#   CAMERA : attendre la fin d'ecriture, puis filmer
# ============================================================

def attendre_fichier_complet(chemin_sans_ext, delai_stable=1.0, timeout=180):
    """Attend que le .cine soit completement ecrit sur le disque.
    On considere l'ecriture finie quand la taille du fichier n'a pas
    bouge pendant 'delai_stable' secondes.
        chemin_sans_ext : chemin SANS l'extension .cine
        delai_stable    : duree (s) de taille constante pour valider
        timeout         : securite, on abandonne au bout de ce delai (s)
    Renvoie True si complet, False si timeout."""
    chemin = chemin_sans_ext + ".cine"
    debut = time.time()
    taille_avant = -1
    instant_dernier_changement = time.time()

    while True:
        # securite : on n'attend jamais indefiniment
        if time.time() - debut > timeout:
            print("  ATTENTION : timeout, fichier pas fini a temps :", chemin)
            return False

        if os.path.exists(chemin):
            taille = os.path.getsize(chemin)
            if taille != taille_avant:
                # la taille change -> l'ecriture continue
                taille_avant = taille
                instant_dernier_changement = time.time()
            elif taille > 0 and (time.time() - instant_dernier_changement) >= delai_stable:
                # taille stable depuis assez longtemps -> c'est fini
                return True

        time.sleep(0.2)   # on re-teste 5 fois par seconde


def filmer(cam, nom_fichier, nb_images, fps):
    """Arme la camera, enregistre, sauvegarde, et ATTEND que le fichier
    soit reellement ecrit sur le disque avant de rendre la main."""
    cam.post_trigger_frames = nb_images
    cam.record()                          # ARME : attend le front du GBF

    # Phase 1 : enregistrement (duree previsible) + marge de securite
    temps_enregistrement = (nb_images / fps) + 2
    print(f"  enregistrement... (~{temps_enregistrement:.1f} s)")
    time.sleep(temps_enregistrement)

    # Phase 2 : sauvegarde sur le disque
    c = cam.Cine(1)
    c.save(
        filename = nom_fichier,
        format   = utils.FileTypeEnum(0),
        range    = utils.FrameRange(c.range.first_image, c.range.last_image)
    )

    # On NE rend la main que quand le fichier est vraiment fini d'ecrire
    if attendre_fichier_complet(nom_fichier):
        print(f"  film bien enregistre : {nom_fichier}.cine")
    else:
        print(f"  PROBLEME : {nom_fichier}.cine semble incomplet")

    # On vide la RAM camera SEULEMENT apres avoir verifie le fichier
    cam.clear_ram()


# ============================================================
#   JOURNAL : recapitulatif Excel des prises
# ============================================================

def ecrire_journal_excel(lignes, chemin_xlsx):
    """Ecrit le journal des prises (une ligne par film) dans un fichier Excel.
    'lignes' est une liste de dictionnaires : chaque dict = une prise, et ses
    cles deviennent les colonnes. Si openpyxl n'est pas installe, on ecrit un
    CSV a la place pour ne rien perdre."""
    if not lignes:
        print("  Journal vide, aucun recapitulatif cree.")
        return

    # Les colonnes = les cles de la 1ere ligne (toutes les lignes ont les memes)
    colonnes = list(lignes[0].keys())

    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        classeur = Workbook()
        feuille = classeur.active
        feuille.title = "Scan PIV"

        feuille.append(colonnes)                       # ligne d'en-tete
        for ligne in lignes:
            feuille.append([ligne[c] for c in colonnes])

        # Largeur des colonnes ajustee au contenu (confort de lecture)
        for i, c in enumerate(colonnes, start=1):
            largeur = len(str(c))
            for ligne in lignes:
                largeur = max(largeur, len(str(ligne[c])))
            feuille.column_dimensions[get_column_letter(i)].width = largeur + 2

        classeur.save(chemin_xlsx)
        print(f"  Recapitulatif Excel ecrit : {chemin_xlsx}")

    except ImportError:
        # openpyxl absent -> on sauve un CSV (Excel l'ouvre tres bien)
        chemin_csv = chemin_xlsx.replace(".xlsx", ".csv")
        with open(chemin_csv, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=colonnes, delimiter=";")
            writer.writeheader()
            writer.writerows(lignes)
        print("  (openpyxl absent : 'pip install openpyxl' pour avoir du .xlsx)")
        print(f"  Recapitulatif CSV ecrit : {chemin_csv}")


# ============================================================
#   UN BALAYAGE EN X DECROISSANT
# ============================================================

def balayage_x(arduino, cam, x_debut, x_fin, pas_entre, nb_images, fps, prefixe):
    """Balaye en X DECROISSANT de x_debut a x_fin, en filmant a chaque pas.
    On suppose qu'on est deja place a x_debut (via aller_a juste avant).
    On filme les deux bornes : la position de depart ET la position finale.
    A chaque film, on ajoute une ligne au journal global."""
    distance = x_debut - x_fin
    if distance <= 0:
        print(f"  ATTENTION : il faut x_debut > x_fin ; balayage '{prefixe}' ignore.")
        return

    nb_positions = distance // pas_entre + 1   # +1 : on filme AUSSI la position finale
    print(f"\n=== Balayage {prefixe} : {nb_positions} positions ===")

    # Reglages camera reellement appliques (constants pendant tout le balayage)
    largeur_px, hauteur_px = cam.resolution
    fps_reel = cam.frame_rate
    expo_reel = cam.exposure

    for i in range(nb_positions):
        print(f"-- {prefixe} position {i + 1}/{nb_positions} --")
        nom = os.path.join(DOSSIER_SORTIE, f"{prefixe}_{i:03d}")

        pos = lire_position(arduino)             # 1. ou est la plateforme MAINTENANT
        filmer(cam, nom, nb_images, fps)         # 2. filmer ici

        # 3. on note tous les parametres de cette prise dans le journal
        journal.append({
            "numero": i,
            "mode": prefixe,
            "fichier": f"{prefixe}_{i:03d}.cine",
            "X (pas)": pos["X"],
            "Y (pas)": pos["Y"],
            "Theta (pas)": pos["T"],
            "resolution (px)": f"{largeur_px}x{hauteur_px}",
            "fps": fps_reel,
            "exposition (us)": expo_reel,
            "nb_images": nb_images,
            "heure": time.strftime("%H:%M:%S"),
        })

        if i < nb_positions - 1:                 # 4. reculer d'un pas SAUF apres le dernier
            bouger(arduino, "XM", pas_entre)


# ============================================================
#   PROGRAMME PRINCIPAL
# ============================================================

os.makedirs(DOSSIER_SORTIE, exist_ok=True)

# Journal : une ligne (un dict) par film. Rempli pendant le scan par
# balayage_x, puis ecrit en Excel a la fin (voir le bloc finally).
journal = []

# --- 1. Reglages camera ---
print("=== Reglages camera (Entree = valeur par defaut) ===")
res_w     = demander_entier("Resolution largeur (px)", 768)
res_h     = demander_entier("Resolution hauteur (px)", 480)
fps       = demander_decimal("Frame rate (images/s)", 90)
expo      = demander_decimal("Exposition (microsecondes)", 8000)
nb_images = demander_entier("Nombre d'images par position", 100)

# --- 2. Reglages mouvements communs ---
print("\n=== Mouvements (Entree = valeur par defaut) ===")
pas_entre = demander_entier("Pas entre 2 images", 15)
pas_90    = demander_entier("Pas pour tourner de 90 degres", 800)

# --- 3. Mode LARGEUR : coordonnees face a la nappe + fin de balayage ---
print("\n=== Mode LARGEUR (coordonnees en pas) ===")
larg_x_debut = demander_entier("X face nappe (debut du scan)", 3250)
larg_y       = demander_entier("Y face camera", 4150)
larg_x_fin   = demander_entier("X de fin du scan", 2905)

# --- 4. Mode LONGUEUR : coordonnees face a la nappe + fin de balayage ---
print("\n=== Mode LONGUEUR (coordonnees en pas) ===")
long_x_debut = demander_entier("X face nappe (debut du scan)", 3400)
long_y       = demander_entier("Y face camera", 4450)
long_x_fin   = demander_entier("X de fin du scan", 2950)

# --- 5. Port serie de l'Arduino ---
print("\n=== Port de l'Arduino ===")
ports = serial.tools.list_ports.comports()
for i, port in enumerate(ports):
    print(i, ":", port)
numero = int(input("Numero du port Arduino : "))
nom_port = ports[numero].device

# --- 6. Connexions ---
arduino = None
ph = None
cam = None

try:
    arduino = serial.Serial(nom_port, 115200, timeout=None)  # None = attend indefiniment
    time.sleep(2)
    print("Plateforme connectee.")

    ph = Phantom()
    cam = camera_selector(ph)
    print("Camera connectee :", cam.model)

    # --- 7. Reglages camera (ordre : resolution -> frequence -> expo) ---
    cam.resolution = (res_w, res_h)
    cam.frame_rate = fps
    cam.exposure = expo
    cam.partition_count = 1

    mode_sync = trouver_mode_sync()
    if mode_sync is not None:
        try:
            cam.sync_mode = mode_sync
        except Exception as e:
            print("  ATTENTION : impossible de regler le mode sync :", e)

    print("Reglages camera appliques :")
    print("  resolution :", cam.resolution)
    print("  frame_rate :", cam.frame_rate, "fps")
    print("  exposure   :", cam.exposure, "us")
    print("  sync_mode  :", cam.sync_mode)

    # ========================================================
    #   LA PIV
    # ========================================================

    # --- Mise a l'origine ---
    print("\nMise a l'origine...")
    envoyer(arduino, "RESET")
    attendre_arrivee(arduino, 0, 0)              # on confirme le retour a l'origine

    # --- MODE LARGEUR ---
    print("\n=== MODE LARGEUR ===")
    aller_a(arduino, larg_x_debut, larg_y)       # va face nappe/camera ET confirme l'arrivee
    balayage_x(arduino, cam, larg_x_debut, larg_x_fin, pas_entre, nb_images, fps, "largeur")

    # --- Rotation de 90 degres ---
    print("\n=== Rotation 90 degres ===")
    bouger(arduino, "TP", pas_90)

    # --- MODE LONGUEUR ---
    print("\n=== MODE LONGUEUR ===")
    aller_a(arduino, long_x_debut, long_y)       # va face nappe/camera ET confirme l'arrivee
    balayage_x(arduino, cam, long_x_debut, long_x_fin, pas_entre, nb_images, fps, "longueur")

    print("\n=== SCAN TERMINE ===")

finally:
    if cam is not None:
        cam.close()
    if ph is not None:
        ph.close()
    if arduino is not None:
        arduino.close()
    print("Materiel deconnecte.")

    # Recapitulatif Excel de toutes les prises.
    # Place dans 'finally' : il est ecrit meme si le scan s'arrete en cours de
    # route (on garde alors la trace des films deja realises). Le nom contient
    # la date/heure pour ne pas ecraser le recapitulatif d'un scan precedent.
    nom_recap = "recap_scan_" + time.strftime("%Y-%m-%d_%Hh%M") + ".xlsx"
    chemin_recap = os.path.join(DOSSIER_SORTIE, nom_recap)
    ecrire_journal_excel(journal, chemin_recap)
